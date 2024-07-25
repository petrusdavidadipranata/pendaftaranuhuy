from flask import Flask, request, render_template, redirect, url_for, session
from flask_bcrypt import Bcrypt
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = 'secret_key'  # Ganti dengan secret key yang lebih aman untuk produksi
bcrypt = Bcrypt(app)

# Contoh data pengguna (sebaiknya diganti dengan database pengguna yang sesungguhnya)
users = {
    'admin': {
        'password': bcrypt.generate_password_hash('st.Martinus2024').decode('utf-8'),  # Ganti dengan password yang di-hash dari database
        'role': 'admin'
    }
}

# Fungsi untuk memeriksa login dengan hashing password
def check_login(username, password):
    if username in users:
        hashed_password = users[username]['password']
        if bcrypt.check_password_hash(hashed_password, password):
            return True, users[username]['role']
    return False, None

# Fungsi untuk membaca data dari file Excel
def read_excel_data():
    data = {}
    counts = {'Total': 0}
    try:
        wb = load_workbook('registrasi.xlsx')
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if row[0] == 'Nama':  # Skip header row
                continue
            entry = dict(zip(['Nama', 'Asal Stasi'], row))
            asal_stasi = entry['Asal Stasi']
            counts['Total'] += 1
            if asal_stasi in data:
                data[asal_stasi].append(entry)
                counts[asal_stasi] += 1
            else:
                data[asal_stasi] = [entry]
                counts[asal_stasi] = 1
    except Exception as e:
        print(f"Error reading Excel file: {str(e)}")
    return data, counts

# Route untuk halaman absensi
@app.route('/absensi')
def absensi():
    if 'logged_in' in session and session['logged_in']:
        try:
            data, counts = read_excel_data()
            return render_template('data_regis.html', data=data, counts=counts)
        except Exception as e:
            print(f"Error: {str(e)}")
            return f'Error: {str(e)}'
    else:
        return redirect(url_for('login'))

@app.route('/succes')
def success():
    return render_template('success.html')
# Route untuk halaman utama (form pendaftaran)
@app.route('/')
def index():
    return render_template('form_regis.html')

# Route untuk memproses data dari form pendaftaran
@app.route('/submit', methods=['POST'])
def submit():
    data = {
        'Nama': request.form['nama'],
        'Asal Stasi': request.form['stasi'],
    }
    try:
        # Simpan ke file Excel
        wb = load_workbook('registrasi.xlsx')
        ws = wb.active
        ws.append([data['Nama'], data['Asal Stasi']])
        wb.save('registrasi.xlsx')
        return '''
        <script>
            window.location.href = "/succes";
        </script>
        '''
    except Exception as e:
        print(f"Error saving to Excel file: {str(e)}")
        return f'Error: {str(e)}'

# Route untuk login
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        authenticated, role = check_login(username, password)
        if authenticated and role == 'admin':
            session['logged_in'] = True
            session['username'] = username
            return redirect(url_for('absensi'))
        else:
            return render_template('login.html', error='Invalid credentials or insufficient permissions. Please try again.')
    return render_template('login.html')

# Route untuk logout
@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    session.pop('username', None)
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(debug=True)
